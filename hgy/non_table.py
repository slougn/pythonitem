# -*- coding: utf-8 -*-
"""
Created on Thu May  5 19:21:48 2022

@author: hangy8
"""

from scipy.optimize import minimize
import pandas as pd
import numpy as np
from scipy import optimize
from openpyxl import load_workbook

BUILDING = pd.read_excel(
    '/home/gree/MyDocument/MyWorkSpace/pythonitem/hgy/data/non_table.xlsx',
    sheet_name='R')

WEATHER = pd.read_excel(
    '/home/gree/MyDocument/MyWorkSpace/pythonitem/hgy/data/non_table.xlsx',
    sheet_name='Tc')

PLANT = pd.read_excel(
    '/home/gree/MyDocument/MyWorkSpace/pythonitem/hgy/data/non_table.xlsx',
    sheet_name='PLANT')

print(PLANT)

CHWP2 = pd.read_excel(
    '/home/gree/MyDocument/MyWorkSpace/pythonitem/hgy/data/non_table.xlsx',
    sheet_name='CHWP2')

TERMINAL = pd.read_excel(
    '/home/gree/MyDocument/MyWorkSpace/pythonitem/hgy/data/non_table.xlsx',
    sheet_name='TERMINAL')

ENERGY = pd.read_excel(
    '/home/gree/MyDocument/MyWorkSpace/pythonitem/hgy/data/non_table.xlsx',
    sheet_name='ENERGY')

#LOAD=pd.read_excel('C://Users//hangy8//Desktop//11111111111111111//non_table//non_table.xlsx',sheet_name='R',usecols=['DATE','LOAD'])
#LOAD['DATE']=pd.to_datetime(LOAD['DATE'])
#LOAD=LOAD.set_index('DATE')
#print(LOAD.resample('w').sum().to_period('w'))
#LOAD_m = LOAD.resample('m').sum().to_period('m')
#print(LOAD_m)
#print(LOAD.resample('Q').sum().to_period('q'))
#print(LOAD.resample('AS').sum().to_period('a'))

LOAD = pd.DataFrame(pd.read_excel('/home/gree/MyDocument/MyWorkSpace/pythonitem/hgy/data/non_table.xlsx',sheet_name='R'))
print(LOAD)
LOAD_M = LOAD.groupby(LOAD['DATE'].apply(lambda x:x.month)).sum()
print(LOAD_M)
LOAD_m1 = LOAD_M.values
LOAD_m = LOAD_m1[:, -1]
print(LOAD_m)

Tc = WEATHER["Tc"].values   #湿球温度
print(Tc)

ENERGY = ENERGY['ENERGY'].values
print(ENERGY)

file_location = '/home/gree/MyDocument/MyWorkSpace/pythonitem/hgy/data/non_table.xlsx'
excel = load_workbook(file_location)
table1 = excel.get_sheet_by_name('PLANT')
table2 = excel.get_sheet_by_name('CHWP2')
table3 = excel.get_sheet_by_name('TERMINAL')

R_CHILLER_1 = table1.cell(row = 2,column = 2).value
R_CHILLER_2 = table1.cell(row = 3,column = 2).value
COP_CHILLER_1 = table1.cell(row = 2,column = 3).value
COP_CHILLER_2 = table1.cell(row = 3,column = 3).value
P_CHILLER_1 = table1.cell(row = 2,column = 4).value
P_CHILLER_2 = table1.cell(row = 3,column = 4).value
print(R_CHILLER_1,R_CHILLER_2,COP_CHILLER_1,COP_CHILLER_2,P_CHILLER_1,P_CHILLER_2)

W_CHWP_1 = table1.cell(row = 2,column = 6).value
W_CHWP_2 = table1.cell(row = 3,column = 6).value
H_CHWP_1 = table1.cell(row = 2,column = 7).value
H_CHWP_2 = table1.cell(row = 3,column = 7).value
P_CHWP_1 = table1.cell(row = 2,column = 8).value
P_CHWP_2 = table1.cell(row = 3,column = 8).value
print(W_CHWP_1,W_CHWP_2,H_CHWP_1,H_CHWP_2,P_CHWP_1,P_CHWP_2)

W_CWP_1 = table1.cell(row = 2,column = 10).value
W_CWP_2 = table1.cell(row = 3,column = 10).value
H_CWP_1 = table1.cell(row = 2,column = 11).value
H_CWP_2 = table1.cell(row = 3,column = 11).value
P_CWP_1 = table1.cell(row = 2,column = 12).value
P_CWP_2 = table1.cell(row = 3,column = 12).value
print(W_CWP_1,W_CWP_2,H_CWP_1,H_CWP_2,P_CWP_1,P_CWP_2)

W_CWT_1 = table1.cell(row = 2,column = 10).value
W_CWT_2 = table1.cell(row = 3,column = 10).value
H_CWT_1 = table1.cell(row = 2,column = 11).value
H_CWT_2 = table1.cell(row = 3,column = 11).value
P_CWT_1 = table1.cell(row = 2,column = 12).value
P_CWT_2 = table1.cell(row = 3,column = 12).value
print(W_CWT_1,W_CWT_2,H_CWT_1,H_CWT_2,P_CWT_1,P_CWT_2)

W_CHWP2_1 = table2.cell(row = 3,column = 2).value
W_CHWP2_2 = table2.cell(row = 4,column = 2).value
H_CHWP2_1 = table2.cell(row = 3,column = 3).value
H_CHWP2_2 = table2.cell(row = 4,column = 3).value
P_CHWP2_1 = table2.cell(row = 3,column = 4).value
P_CHWP2_2 = table2.cell(row = 4,column = 4).value
print(W_CHWP2_1,W_CHWP2_2,H_CHWP2_1,H_CHWP2_2,P_CHWP2_1,P_CHWP2_2)

P_PerR = table3.cell(row = 3,column = 2).value
print(P_PerR)

Coe_Energy_m = ENERGY/max(ENERGY)
print(Coe_Energy_m)

Coe_Load_m = LOAD_m/max(LOAD_m)
print(Coe_Load_m)

Coe_EnergyLoad_m = Coe_Energy_m / Coe_Load_m
print(Coe_EnergyLoad_m)
Coe_EnergyLoad_m_list = Coe_EnergyLoad_m.tolist()
print(Coe_EnergyLoad_m_list)

LOAD1 = LOAD[(LOAD['DATE']>="1999-01-01") & (LOAD['DATE']<"1999-02-01")]
LOAD2 = LOAD[(LOAD['DATE']>="1999-02-01") & (LOAD['DATE']<"1999-03-01")]
LOAD3 = LOAD[(LOAD['DATE']>="1999-03-01") & (LOAD['DATE']<"1999-04-01")]
LOAD4 = LOAD[(LOAD['DATE']>="1999-04-01") & (LOAD['DATE']<"1999-05-01")]
LOAD5 = LOAD[(LOAD['DATE']>="1999-05-01") & (LOAD['DATE']<"1999-06-01")]
LOAD6 = LOAD[(LOAD['DATE']>="1999-06-01") & (LOAD['DATE']<"1999-07-01")]
LOAD7 = LOAD[(LOAD['DATE']>="1999-07-01") & (LOAD['DATE']<"1999-08-01")]
LOAD8 = LOAD[(LOAD['DATE']>="1999-08-01") & (LOAD['DATE']<"1999-09-01")]
LOAD9 = LOAD[(LOAD['DATE']>="1999-09-01") & (LOAD['DATE']<"1999-10-01")]
LOAD10 = LOAD[(LOAD['DATE']>="1999-10-01") & (LOAD['DATE']<"1999-11-01")]
LOAD11 = LOAD[(LOAD['DATE']>="1999-11-01") & (LOAD['DATE']<"1999-12-01")]
LOAD12 = LOAD[(LOAD['DATE']>="1999-12-01") & (LOAD['DATE']<="1999-12-31")]

print(LOAD1)
print(LOAD2)
print(LOAD3)
print(LOAD4)
print(LOAD5)
print(LOAD6)
print(LOAD7)
print(LOAD8)
print(LOAD9)
print(LOAD10)
print(LOAD11)
print(LOAD12)

LOAD1_list0 = LOAD1['LOAD'].tolist()
LOAD2_list0 = LOAD2['LOAD'].tolist()
LOAD3_list0 = LOAD3['LOAD'].tolist()
LOAD4_list0 = LOAD4['LOAD'].tolist()
LOAD5_list0 = LOAD5['LOAD'].tolist()
LOAD6_list0 = LOAD6['LOAD'].tolist()
LOAD7_list0 = LOAD7['LOAD'].tolist()
LOAD8_list0 = LOAD8['LOAD'].tolist()
LOAD9_list0 = LOAD9['LOAD'].tolist()
LOAD10_list0 = LOAD10['LOAD'].tolist()
LOAD11_list0 = LOAD11['LOAD'].tolist()
LOAD12_list0 = LOAD12['LOAD'].tolist()

print(LOAD1_list0)
print(LOAD2_list0)
print(LOAD3_list0)
print(LOAD4_list0)
print(LOAD5_list0)
print(LOAD6_list0)
print(LOAD7_list0)
print(LOAD8_list0)
print(LOAD9_list0)
print(LOAD10_list0)
print(LOAD11_list0)
print(LOAD12_list0)

#冷负荷
LOAD1_list1 = (np.array(LOAD1_list0) * Coe_EnergyLoad_m_list[0]).tolist()  
LOAD2_list1 = (np.array(LOAD2_list0) * Coe_EnergyLoad_m_list[1]).tolist()
LOAD3_list1 = (np.array(LOAD3_list0) * Coe_EnergyLoad_m_list[2]).tolist()
LOAD4_list1 = (np.array(LOAD4_list0) * Coe_EnergyLoad_m_list[3]).tolist()
LOAD5_list1 = (np.array(LOAD5_list0) * Coe_EnergyLoad_m_list[4]).tolist()
LOAD6_list1 = (np.array(LOAD6_list0) * Coe_EnergyLoad_m_list[5]).tolist()
LOAD7_list1 = (np.array(LOAD7_list0) * Coe_EnergyLoad_m_list[6]).tolist()
LOAD8_list1 = (np.array(LOAD8_list0) * Coe_EnergyLoad_m_list[7]).tolist()
LOAD9_list1 = (np.array(LOAD9_list0) * Coe_EnergyLoad_m_list[8]).tolist()
LOAD10_list1 = (np.array(LOAD10_list0) * Coe_EnergyLoad_m_list[9]).tolist()
LOAD11_list1 = (np.array(LOAD11_list0) * Coe_EnergyLoad_m_list[10]).tolist()
LOAD12_list1 = (np.array(LOAD12_list0) * Coe_EnergyLoad_m_list[11]).tolist()
#print(LOAD1_list1)
I = LOAD1_list1+LOAD2_list1+LOAD3_list1+LOAD4_list1+LOAD5_list1+LOAD6_list1+LOAD7_list1+LOAD8_list1+LOAD9_list1+LOAD10_list1+LOAD11_list1+LOAD12_list1

#修正的冷负荷
param = np.ones(12)
def LOAD_fix(I,param):  #函数未写完
    J = []
    #         1  2  3  4  5  6  7  8  9  10  11  12
    month = [31, 28,31,30,31,30,31,31,30,31, 30, 31]
    for j in range(len(I)):
        if j <= month[0] * 24:
            J.append(I[j]*param[0])
        else:
            J.append(0)
    return J

J = LOAD_fix(I,param)


# 冷却水进水温度
K = Tc

T_in = Tc + 5
for i in range(len(T_in)):
    if T_in[i] < 18:
        T_in[i] = 18
    else:
        continue

L = T_in

# 运行台数
def machine_count(J3):
    N = []
    for i in range(len(J3)):
        if J3[i] == 0:
            N.append(0)
        else:
            N.append(round(J3[i]/350/3.517))
    return N

N = machine_count(J)

# 冷水机组1号 运行台数
def cool_machine1_count(N3):
    O = []
    for i in range(len(N3)):
        if N3[i] >= 1:
            O.append(1)
        else:
            O.append(0)
    return O

O = cool_machine1_count(N)
# 冷水机组2号 运行台数
def cool_machine2_count(N3):
    P = []
    for i in range(len(N3)):
        if N3[i] >= 2:
            P.append(1)
        else:
            P.append(0)
    return P

P = cool_machine2_count(N)
# 冷水机组1号 负荷率
def cool_machine1_load(O3,J3,P3):
    Q = []
    for i in range(len(O3)):
        if O3[i] == 0:
            Q.append(0)
        else:
            Q.append(J3[i]/(350*3.517*O3[i]+350*3.517*P3[i]))
    return Q

Q = cool_machine1_load(O,J,P)
# 冷水机组2号 负荷率
def cool_machine2_load(P3,J3,O3):
    R = []
    for i in range(len(P3)):
        if P3[i] == 0:
            R.append(0)
        else:
            R.append(J3/(350*3.517*O3[i]+350*3.517*P3[i]))
    return R

R = cool_machine2_load(P,J,O)

# 冷冻出水温度
T_out = (np.ones(len(T_in))*7.5).tolist()

S = T_out

# 冷水机组1号修正负荷率
def cool_machine1_load_fix(Q3):
    #=IF(AND(Q3<0.1),0,IF(AND(Q3>=0.1,Q3<0.3),0.3,IF(AND(Q3>1),1,Q3)))
    T = []
    for i in range(len(Q3)):
        if Q3[i] < 0.1 :
            T.append(0)
        elif Q3[i] >= 0.1 and Q3[i] < 0.3:
            T.append(0.3)
        elif Q3[i] >1 :
            T.append(1)
        else:
            T.append(Q3[i])
    return T

T = cool_machine1_load_fix(Q)

# 冷水机组2号修正负荷率
def cool_machine2_load_fix(R3):
    U = cool_machine1_load_fix(R3)
    return U


# AB列公式  COP公式
def COP(X1,X2,X3):
    return (np.exp(0.000407461117728266 * X1 + 0.0319727844024694 * X2 + -0.0321549861357119 * X3 + 2.49432757221509)).tolist()

# 1号COP
def machine1_COP(T3,S3,M3):
    V = []
    for i in range(len(T3)):
        if T3[i] == 0:
            V.append(0)
        else:
            V.append(COP(T3*100,S3,M3)*(1-0.02*10)*0.9)
    return V

# 2号COP
def machine2_COP(U3,S3,M3):
    W = machine1_COP(U3,S3,M3)
    return W

# 1号制冷量
def machine1_cool_capacity(T3):
    X_cool_capacity = (350*3.517*np.array(T3)).tolist()
    return X_cool_capacity

# 2号制冷量
def machine2_cool_capacity(U3):
    Y_cool_capacity = (350*3.517*np.array(U3)).tolist()
    return Y_cool_capacity

# 1号功率
def machine1_power(V3,X3):
    Z = []
    for i in range(len(V3)):
        if V3[i] == 0:
            Z.append(0)
        else:
            Z.append(X3[i]/V3[i])
    return Z

# 2号功率
def machine2_power(W3,Y3):
    AA = machine1_power(W3,Y3)
    return AA

# 1号冷冻水流量
def machine1_flow(X3):
    AG = []
    for i in range(len(X3)):
        AG.append(X3[i]*0.172*5/3)
    return AG

# 2号冷冻水流量
def machine2_flow(Y3):
    AH = machine1_flow(Y3)
    return AH

# 1号冷冻水泵频率
def cool_machine1_Hz(AG3):
    AI = []
    for i in range(len(AG3)):
        if AG3[i]/250 == 0:
            AI.append(0)
        elif 0 < AG3[i]/250 * 50 < 30:
            AI.append(30)
        elif AG3[i]/250 * 50 > 50:
            AI.append(50)
        else:
            AI.append(AG3[i]/250 * 50)
    return AI

# 2号冷冻水泵频率
def cool_machine2_Hz(AH3):
    AJ = cool_machine1_Hz(AH3)
    return AJ

# 1号冷冻水泵功率
def cool_machine1_power(AI3):
    AK = []
    for i in range(len(AI3)):
        if 55*(AI3[i]/50)**3/0.85/0.9 > 55:
            AK.append(55)
        else:
            AK.append(55*(AI3[i]/50)**3/0.85/0.9)
    return AK

# 2好冷冻水泵功率
def cool_machine2_power(AJ3):
    AL = cool_machine1_power(AJ3)
    return AL

# 冷冻水泵功率
def cool_machine_power(AK,AL):
    AM = []
    for i in range(len(AK)):
        AM.append(AK[i] + AL[i])
    return AM

# 1号冷却水流量
def lengque1_flow(X3):
    AN = []
    for i in range(len(X3)):
        AN.append(X3[i]*0.125*5/3)
    return AN

# 2号冷却水流量
def lengque2_flow(Y3):
    AO = lengque1_flow(Y3)
    return AO

# 1号冷却水泵频率
def lengque1_Hz(AN3):
    AP = []
    for i in range(len(AN3)):
        if AN3[i]/300 == 0:
            AP.append(0)
        elif 0 < AN3[i]/300 * 50 < 30:
            AP.append(30)
        elif AN3[i]/250 * 50 > 50:
            AP.append(50)
        else:
            AP.append(AN3[i]/250 * 50)
    return AP

# 2号冷却水泵频率
def lengque2_Hz(AO3):
    AQ = lengque1_Hz(AO3)
    return AQ

# 1号冷却水泵功率
def lengque1_power(AP3):
    AR = cool_machine1_power(AP3)
    return AR

# 2号冷却水泵功率
def lengque2_power(AQ3):
    AS = lengque1_power(AQ3)
    return AS

# 冷却水泵功率
def lengque_power(AR,AS):
    AT = []
    for i in range(len(AR)):
        AT.append(AR[i] + AS[i])
    return AT

# 1号冷却塔流量修正
def lengque_tower1_fix(AN3):
    AU = []
    for i in range(len(AN3)):
        if AN3[i] == 0:
            AU.append(0)
        elif 0 < AN3[i] < 300*25/50:
            AU.append(300*25/50)
        else:
            AU.append(AN3[i])
    return AU

# 2号冷却塔流量修正
def lengque_tower2_fix(AO3):
    AV = lengque_tower1_fix(AO3)
    return AV

# 1号冷却塔功率
def lengque_tower1_power(AU3):
    #=IF(22*(AU3/300)^3/0.85>22,22,22*(AU3/300)^3/0.85)
    AW = []
    for i in range(len(AU3)):
        if 22*(AU3[i]/300)**3/0.85 > 22:
            AW.append(22)
        else:
            AW.append(22*(AU3[i]/300)**3/0.85 )
    return AW

# 2号冷却塔功率
def lengque_tower2_power(AV3):
    AX = lengque_tower1_power(AV3)
    return AX

# 冷却塔功率 
def lengque_tower_power(AW,AX):
    AY = []
    for i in range(len(AW)):
        AY.append(AW[i]+ AX[i])
    return AY

# 二级功率
def secondary_power(AG3):
    AZ = []
    for i in range(len(AG3)):
        if AG3[i] == 0:
            AZ.append(0)
        else:
            AZ.append(0.75)

# 末端功率
def terminal_power(AZ3,X3,Y3):
    BA = []
    for i in range(len(AZ3)):
        if AZ3[i] == 0:
            BA.append(0)
        else:
            BA.append(0.15*(X3[i]+Y3[i]))
    return BA

# 空调系统功率
def air_conditioner_power(AD3,AM3,AT3,AY3,AZ3,BA3):
    BB = []
    for i in range(len(AD3)):
        BB.append(AD3[i]+AM3[i]+AT3[i]+AY3[i]+AZ3[i]+BA3[i])
    return BB



#LOAD_1 = LOAD * Coe_EnergyLoad_m
#print(LOAD_1)

#LOAD_2 = LOAD_1 * param
    
    
